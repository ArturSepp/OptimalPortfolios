"""Paper-local replay of the engine used for the August 2024 update.

The shared :mod:`optimalportfolios` package has deliberately evolved since the
paper update.  In particular, the ERC solver, the MaxSharpe estimators and the
Gaussian-mixture implementation are no longer the algorithms that produced the
archived workbook.  This module freezes those numerical contracts at paper
scope.  It does not fetch data and it does not modify the production package.

The reference source is git commit
``6038fba806b381dfb6af2c7e75a8002b57ee66e7`` (v2.1.1, 20 August 2024).
The historical dependency file contained lower bounds rather than a lockfile,
so the runtime report below records the actual replay environment and states
explicitly that bitwise reconstruction of the original environment is not
possible from the repository alone.
"""

from __future__ import annotations

import hashlib
import importlib.metadata
import platform
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Callable, Iterable, Mapping, Sequence

import cvxpy as cvx
import numpy as np
import pandas as pd
import qis
from packaging.version import InvalidVersion, Version
from scipy.optimize import minimize

import optimalportfolios as opt


REFERENCE_COMMIT = "6038fba806b381dfb6af2c7e75a8002b57ee66e7"
REFERENCE_PACKAGE_VERSION = "2.1.1"

# These are the lower bounds present in the 2024 project requirements.  They
# are compatibility checks, not claims about the exact versions used to create
# the workbook: no historical lockfile was committed.
HISTORICAL_DEPENDENCY_FLOORS: Mapping[str, str] = {
    "numpy": "1.26.4",
    "pandas": "2.2.0",
    "scipy": "1.12.0",
    "qis": "2.1.1",
    "cvxpy": "1.3.2",
    "scikit-learn": "1.3",
}

GOLDEN_WORKBOOK_PATH = Path(
    r"C:\Users\artur\OneDrive\analytics\outputs\Crypto_portfolios_backtests_20240820_1532.pdf.xlsx"
)
GOLDEN_WORKBOOK_SHA256 = (
    "d19e3a5ed62409287a4673c74e29d65bccbaf5df7380acf99bef20f86eee1176"
)
GOLDEN_PANEL_PATH = (
    Path(__file__).resolve().parent.parent / "data" / "crypto_allocation_prices_updated.csv"
)

# Workbook row 3 is "Median crypto weight".  Columns B:E are ERC, MaxDiv,
# MaxSharpe and CARA-3; F is their cross-method median.
GOLDEN_MEDIAN_WEIGHT_ROWS: Mapping[str, tuple[float, float, float, float, float]] = {
    "weight_(A) 100% Alts with BTC": (
        0.01745299827073789,
        0.02188897347653096,
        0.09209724211575182,
        0.2142456776432366,
        0.05699310779614139,
    ),
    "weight_(C) 75%25% BalAlts with BTC": (
        0.005773754497725937,
        0.03126069596751974,
        0.04477507459839808,
        0.1924499273909754,
        0.03801788528295891,
    ),
    "weight_(B) 100% Alts with ETH": (
        0.01326709547380079,
        0.01905670401524881,
        0.04127258023333454,
        0.08078610175262163,
        0.03016464212429167,
    ),
    "weight_(D) 75%25% BalAlts with ETH": (
        0.004876334910805515,
        0.03044928169629908,
        0.01785774681844323,
        0.08376252135308825,
        0.02415351425737115,
    ),
}
GOLDEN_HEADLINE_MEDIAN = 0.03409126


class ParityMethod(str, Enum):
    """The four optimisation methods reported in the paper update."""

    ERC = "ERC"
    MAX_DIV = "MaxDiv"
    MAX_SHARPE = "MaxSharpe"
    CARA = "CARA-3"


class ParityDependencyError(RuntimeError):
    """Raised when a requested historical numerical dependency is unavailable."""


class ParitySolveError(RuntimeError):
    """Raised when a paper-era optimiser does not return admissible weights."""


@dataclass(frozen=True)
class Parity2024Config:
    """Parameters used by the August 2024 paper update."""

    returns_freq: str = "ME"
    rebalancing_freq: str = "QE"
    roll_window: int = 60
    span: int = 30
    carra: float = 0.5
    n_components: int = 3
    first_asset_target_weight: float = 0.75
    rebalancing_costs: float = 0.005
    weight_implementation_lag: int = 1
    estimation_start: str | pd.Timestamp | None = "2010-07-19"
    reporting_start: str | pd.Timestamp | None = "2016-03-31"
    end_date: str | pd.Timestamp | None = None
    max_sharpe_solver: str = "ECOS_BB"


@dataclass(frozen=True)
class DependencyReport:
    """Installed versions and historical-replay readiness."""

    python: str
    packages: Mapping[str, str | None]
    installed_cvxpy_solvers: tuple[str, ...]
    missing_or_incompatible: tuple[str, ...]
    exact_historical_versions_known: bool = False

    @property
    def ready(self) -> bool:
        """Return whether every requested compatibility dependency is present."""

        return not self.missing_or_incompatible

    def to_dict(self) -> dict[str, object]:
        """Return a JSON-serialisable runtime manifest fragment."""

        return {
            "reference_commit": REFERENCE_COMMIT,
            "reference_package_version": REFERENCE_PACKAGE_VERSION,
            "python": self.python,
            "packages": dict(self.packages),
            "installed_cvxpy_solvers": list(self.installed_cvxpy_solvers),
            "missing_or_incompatible": list(self.missing_or_incompatible),
            "ready": self.ready,
            "exact_historical_versions_known": self.exact_historical_versions_known,
            "environment_note": (
                "The 2024 requirements specified lower bounds and supplied no lockfile; "
                "the archived workbook is the numerical oracle."
            ),
        }


@dataclass(frozen=True)
class RollingSharpeInputs:
    """Point-in-time 60-month means and covariance matrices for MaxSharpe."""

    expected_returns: pd.DataFrame
    covariances: Mapping[pd.Timestamp, pd.DataFrame]
    raw_covariances: Mapping[pd.Timestamp, pd.DataFrame]
    window_bounds: pd.DataFrame


@dataclass(frozen=True)
class MixtureParameters:
    """Annualised Gaussian-mixture inputs to the CARA optimiser."""

    means: tuple[np.ndarray, ...]
    covariances: tuple[np.ndarray, ...]
    probabilities: np.ndarray


@dataclass(frozen=True)
class MarginalWeights:
    """Target-weight panels for universes without and with the marginal asset."""

    without_asset: pd.DataFrame
    with_asset: pd.DataFrame


@dataclass(frozen=True)
class MarginalPortfolios:
    """Backtested portfolios for universes without and with the marginal asset."""

    without_asset: qis.PortfolioData
    with_asset: qis.PortfolioData
    weights: MarginalWeights


MixtureFitter = Callable[[np.ndarray, int, float], MixtureParameters]


def _coerce_method(method: ParityMethod | str | object) -> ParityMethod:
    """Normalise paper and current-package method labels."""

    if isinstance(method, ParityMethod):
        return method
    value = getattr(method, "value", method)
    aliases = {
        "ERC": ParityMethod.ERC,
        "MaxDiv": ParityMethod.MAX_DIV,
        "MAX_DIV": ParityMethod.MAX_DIV,
        "MaxSharpe": ParityMethod.MAX_SHARPE,
        "MAX_SHARPE": ParityMethod.MAX_SHARPE,
        "CARA-3": ParityMethod.CARA,
        "MIXTURE": ParityMethod.CARA,
        "CARA": ParityMethod.CARA,
    }
    try:
        return aliases[str(value)]
    except KeyError as exc:
        raise ValueError(f"Unsupported 2024 parity method: {value!r}") from exc


def _installed_version(distribution: str) -> str | None:
    """Return an installed distribution version without importing it."""

    try:
        return importlib.metadata.version(distribution)
    except importlib.metadata.PackageNotFoundError:
        return None


def dependency_report(
    methods: Iterable[ParityMethod | str | object] = (),
) -> DependencyReport:
    """Inspect lower bounds plus method-specific sklearn/ECOS_BB requirements.

    The base numerical stack is always checked.  ``scikit-learn`` is required
    only for CARA and the historical ``ECOS_BB`` backend only for MaxSharpe.
    Installed versions are reported even when a method does not require them so
    every numerical replay can persist the same audit manifest.
    """

    normalised = {_coerce_method(method) for method in methods}
    packages = {
        package: _installed_version(package)
        for package in HISTORICAL_DEPENDENCY_FLOORS
    }
    packages["optimalportfolios"] = _installed_version("optimalportfolios")
    missing: list[str] = []
    base_packages = ("numpy", "pandas", "scipy", "qis", "cvxpy")
    required_packages = list(base_packages)
    if ParityMethod.CARA in normalised:
        required_packages.append("scikit-learn")

    for package in required_packages:
        installed = packages[package]
        floor = HISTORICAL_DEPENDENCY_FLOORS[package]
        if installed is None:
            missing.append(f"{package}>={floor} is not installed")
            continue
        try:
            if Version(installed) < Version(floor):
                missing.append(f"{package}=={installed} is below the historical floor {floor}")
        except InvalidVersion:
            missing.append(f"{package} has an unparseable version {installed!r}")

    installed_solvers = tuple(sorted(cvx.installed_solvers()))
    if ParityMethod.MAX_SHARPE in normalised and "ECOS_BB" not in installed_solvers:
        missing.append("CVXPY solver ECOS_BB is not installed")

    return DependencyReport(
        python=platform.python_version(),
        packages=packages,
        installed_cvxpy_solvers=installed_solvers,
        missing_or_incompatible=tuple(missing),
    )


def require_parity_dependencies(
    methods: Iterable[ParityMethod | str | object],
) -> DependencyReport:
    """Return the runtime report or raise before a non-parity replay can start."""

    report = dependency_report(methods=methods)
    if not report.ready:
        details = "; ".join(report.missing_or_incompatible)
        raise ParityDependencyError(f"2024 parity dependencies are not ready: {details}")
    return report


def _require_cvxpy_solver(solver: str) -> None:
    """Require an explicitly requested solver; never substitute silently."""

    if solver not in cvx.installed_solvers():
        raise ParityDependencyError(
            f"CVXPY solver {solver!r} is not installed; available solvers are "
            f"{sorted(cvx.installed_solvers())}. The historical solver was 'ECOS_BB'."
        )


def _timestamp(value: str | pd.Timestamp | None) -> pd.Timestamp | None:
    """Normalise an optional date boundary."""

    return None if value is None else pd.Timestamp(value)


def _validate_prices(prices: pd.DataFrame) -> pd.DataFrame:
    """Validate and date-truncate an input price panel without filling it."""

    if not isinstance(prices, pd.DataFrame):
        raise TypeError("prices must be a pandas DataFrame")
    if not isinstance(prices.index, pd.DatetimeIndex):
        raise TypeError("prices must have a DatetimeIndex")
    if prices.empty or prices.shape[1] < 2:
        raise ValueError("prices must contain at least two assets")
    if prices.columns.has_duplicates:
        raise ValueError("price columns must be unique")
    if prices.index.has_duplicates or not prices.index.is_monotonic_increasing:
        raise ValueError("price dates must be unique and increasing")
    finite = prices.to_numpy(dtype=float)
    if np.isinf(finite).any():
        raise ValueError("prices must not contain infinite observations")
    if np.any(finite[np.isfinite(finite)] <= 0.0):
        raise ValueError("finite price observations must be strictly positive")
    return prices.astype(float)


def _prices_through_end(prices: pd.DataFrame, config: Parity2024Config) -> pd.DataFrame:
    """Remove observations after the explicit replay cutoff."""

    prices = _validate_prices(prices)
    end = _timestamp(config.end_date)
    return prices if end is None else prices.loc[:end]


def monthly_log_returns_2024(
    prices: pd.DataFrame,
    config: Parity2024Config = Parity2024Config(),
) -> pd.DataFrame:
    """Compute the paper engine's monthly log-return panel."""

    prices = _prices_through_end(prices, config)
    returns = qis.to_returns(
        prices=prices,
        is_log_returns=True,
        drop_first=True,
        freq=config.returns_freq,
    )
    if not isinstance(returns, pd.DataFrame):
        raise TypeError("qis.to_returns did not return a DataFrame")
    return returns


def rolling_return_windows_2024(
    prices: pd.DataFrame,
    config: Parity2024Config = Parity2024Config(),
) -> Mapping[pd.Timestamp, pd.DataFrame]:
    """Return exact trailing observation windows on quarterly rebalance dates."""

    returns = monthly_log_returns_2024(prices=prices, config=config)
    schedule = qis.generate_rebalancing_indicators(
        df=returns,
        freq=config.rebalancing_freq,
    )
    windows: dict[pd.Timestamp, pd.DataFrame] = {}
    for idx, (date, rebalance) in enumerate(schedule.items()):
        if idx >= config.roll_window - 1 and bool(rebalance):
            windows[pd.Timestamp(date)] = returns.iloc[
                idx - config.roll_window + 1 : idx + 1
            ]
    return windows


def estimate_ewm_covariances_2024(
    prices: pd.DataFrame,
    config: Parity2024Config = Parity2024Config(),
) -> Mapping[pd.Timestamp, pd.DataFrame]:
    """Estimate the expanding EWMA covariance sequence used by ERC and MaxDiv."""

    returns = monthly_log_returns_2024(prices=prices, config=config)
    values = returns.to_numpy()
    centred = values - qis.compute_ewm(values, span=config.span)
    tensor = qis.compute_ewm_covar_tensor(
        a=centred,
        span=config.span,
        nan_backfill=qis.NanBackfill.ZERO_FILL,
    )
    annualisation = qis.get_annualization_factor(freq=config.returns_freq)
    schedule = qis.generate_rebalancing_indicators(
        df=returns,
        freq=config.rebalancing_freq,
    )
    start = _timestamp(config.estimation_start)
    end = _timestamp(config.end_date)
    covariances: dict[pd.Timestamp, pd.DataFrame] = {}
    for idx, (date, rebalance) in enumerate(schedule.items()):
        date = pd.Timestamp(date)
        if not bool(rebalance):
            continue
        if start is not None and date < start:
            continue
        if end is not None and date > end:
            continue
        covariances[date] = pd.DataFrame(
            annualisation * tensor[idx],
            index=returns.columns,
            columns=returns.columns,
        )
    return covariances


def estimate_max_sharpe_inputs_2024(
    prices: pd.DataFrame,
    config: Parity2024Config = Parity2024Config(),
) -> RollingSharpeInputs:
    """Reconstruct the paper's trailing-60-month mean and seeded EWMA covariance.

    The unregularised covariance from one quarterly solve seeds the next one.
    The seed is updated *before* the current covariance is regularised, matching
    commit ``6038fba`` exactly.  Means are arithmetic averages of log returns;
    no risk-free rate is subtracted inside the optimiser.
    """

    returns = monthly_log_returns_2024(prices=prices, config=config)
    schedule = qis.generate_rebalancing_indicators(
        df=returns,
        freq=config.rebalancing_freq,
    )
    annualisation = qis.get_annualization_factor(freq=config.returns_freq)
    start = _timestamp(config.estimation_start)
    end = _timestamp(config.end_date)
    covar0 = np.zeros((returns.shape[1], returns.shape[1]))
    means: dict[pd.Timestamp, pd.Series] = {}
    covariances: dict[pd.Timestamp, pd.DataFrame] = {}
    raw_covariances: dict[pd.Timestamp, pd.DataFrame] = {}
    bounds: dict[pd.Timestamp, dict[str, object]] = {}

    for idx, (date, rebalance) in enumerate(schedule.items()):
        if idx < config.roll_window - 1 or not bool(rebalance):
            continue
        date = pd.Timestamp(date)
        window = returns.iloc[idx - config.roll_window + 1 : idx + 1]
        values = window.to_numpy()
        mean = annualisation * np.nanmean(values, axis=0)
        raw_covar = qis.compute_ewm_covar(
            a=values,
            span=config.span,
            covar0=covar0,
        )
        covar0 = np.array(raw_covar, copy=True)
        regularised = qis.matrix_regularization(covar=raw_covar, cut=1e-5)

        # Historical estimation is sequential from the beginning of the panel;
        # pre-start covariances must still update covar0 even when not returned.
        if start is not None and date < start:
            continue
        if end is not None and date > end:
            continue
        means[date] = pd.Series(mean, index=returns.columns)
        raw_covariances[date] = pd.DataFrame(
            raw_covar,
            index=returns.columns,
            columns=returns.columns,
        )
        covariances[date] = pd.DataFrame(
            annualisation * regularised,
            index=returns.columns,
            columns=returns.columns,
        )
        bounds[date] = {
            "first_observation": window.index[0],
            "last_observation": window.index[-1],
            "observations": len(window),
        }

    expected_returns = pd.DataFrame.from_dict(means, orient="index")
    expected_returns = expected_returns.reindex(columns=returns.columns)
    window_bounds = pd.DataFrame.from_dict(bounds, orient="index")
    return RollingSharpeInputs(
        expected_returns=expected_returns,
        covariances=covariances,
        raw_covariances=raw_covariances,
        window_bounds=window_bounds,
    )


def fit_sklearn_mixture_2024(
    values: np.ndarray,
    n_components: int,
    annualisation: float,
) -> MixtureParameters:
    """Fit the exact sklearn GaussianMixture specification used in 2024."""

    report = dependency_report(methods=(ParityMethod.CARA,))
    if not report.ready:
        details = "; ".join(report.missing_or_incompatible)
        raise ParityDependencyError(f"CARA-3 parity is unavailable: {details}")
    from sklearn.mixture import GaussianMixture

    model = GaussianMixture(
        n_components=n_components,
        covariance_type="full",
        random_state=3,
    )
    model.fit(values)
    return MixtureParameters(
        means=tuple(annualisation * value for value in model.means_),
        covariances=tuple(annualisation * value for value in model.covariances_),
        probabilities=np.asarray(model.weights_, dtype=float),
    )


def _legacy_scipy_constraints(
    min_weights: np.ndarray | None,
    max_weights: np.ndarray | None,
) -> list[dict[str, object]]:
    """Build the simple long-only/full-investment constraints from v2.1.1."""

    # Current Constraints.set_scipy_constraints is intentionally rejected here:
    # it translates box limits to SciPy bounds and adds current validation, while
    # v2.1.1 represented them as inequality functions.  The distinction changes
    # SLSQP's numerical path and is material to the archived paper weights.
    constraints: list[dict[str, object]] = [
        {"type": "ineq", "fun": lambda weights: weights},
        {"type": "eq", "fun": lambda weights: 1.0 - np.sum(weights)},
    ]
    if min_weights is not None:
        minimum = np.asarray(min_weights, dtype=float)
        constraints.append(
            {"type": "ineq", "fun": lambda weights, floor=minimum: weights - floor}
        )
    if max_weights is not None:
        maximum = np.asarray(max_weights, dtype=float)
        constraints.append(
            {"type": "ineq", "fun": lambda weights, cap=maximum: cap - weights}
        )
    return constraints


def _validate_solution(
    weights: np.ndarray,
    min_weights: np.ndarray | None,
    max_weights: np.ndarray | None,
    context: str,
    tolerance: float = 2e-5,
) -> np.ndarray:
    """Fail loudly if a historical solver result is unusable."""

    weights = np.asarray(weights, dtype=float)
    if weights.ndim != 1 or not np.all(np.isfinite(weights)):
        raise ParitySolveError(f"{context}: solver returned non-finite weights")
    if abs(float(np.sum(weights)) - 1.0) > tolerance:
        raise ParitySolveError(f"{context}: weights do not sum to one")
    if np.min(weights) < -tolerance:
        raise ParitySolveError(f"{context}: long-only constraint was violated")
    if min_weights is not None and np.any(weights < min_weights - tolerance):
        raise ParitySolveError(f"{context}: minimum-weight constraint was violated")
    if max_weights is not None and np.any(weights > max_weights + tolerance):
        raise ParitySolveError(f"{context}: maximum-weight constraint was violated")
    return weights


def _risk_budget_objective(weights: np.ndarray, parameters: Sequence[np.ndarray]) -> float:
    """Historical sum-of-squares risk-budget objective."""

    covar, budget = parameters
    contributions = qis.compute_portfolio_risk_contributions(w=weights, covar=covar)
    portfolio_vol = np.sqrt(float(weights.T @ covar @ weights))
    if budget is None:
        target = portfolio_vol * np.ones_like(contributions) / len(contributions)
    else:
        target = np.where(
            np.isnan(budget),
            contributions,
            portfolio_vol * budget,
        )
    return float(np.nansum(np.square(contributions - target)))


def _solve_erc_2024(covar: np.ndarray, risk_budget: np.ndarray, context: str) -> np.ndarray:
    """Solve one paper-era ERC allocation with the original SLSQP settings."""

    n_assets = covar.shape[0]
    result = minimize(
        _risk_budget_objective,
        np.ones(n_assets) / n_assets,
        args=([covar, risk_budget],),
        method="SLSQP",
        constraints=_legacy_scipy_constraints(None, None),
        options={"ftol": 1e-18, "maxiter": 200},
    )
    return _validate_solution(result.x, None, None, context=context)


def _max_diversification_objective(
    weights: np.ndarray,
    parameters: Sequence[np.ndarray],
) -> float:
    """Return the negative diversification ratio through the public stack API."""

    return -float(opt.calculate_diversification_ratio(w=weights, covar=parameters[0]))


def _solve_max_div_2024(
    covar: np.ndarray,
    min_weights: np.ndarray | None,
    max_weights: np.ndarray | None,
    context: str,
) -> np.ndarray:
    """Solve one paper-era maximum-diversification allocation."""

    n_assets = covar.shape[0]
    result = minimize(
        _max_diversification_objective,
        np.ones(n_assets) / n_assets,
        args=([covar],),
        method="SLSQP",
        constraints=_legacy_scipy_constraints(min_weights, max_weights),
        options={"disp": False, "ftol": 1e-18, "maxiter": 200},
    )
    # v2.1.1 clipped SLSQP's tiny negative MaxDiv weights, without renormalising.
    weights = np.where(result.x > 0.0, result.x, 0.0)
    return _validate_solution(weights, min_weights, max_weights, context=context)


def _solve_max_sharpe_2024(
    covar: np.ndarray,
    means: np.ndarray,
    min_weights: np.ndarray | None,
    max_weights: np.ndarray | None,
    solver: str,
    context: str,
) -> np.ndarray:
    """Solve the original Charnes-Cooper program with an explicit backend."""

    _require_cvxpy_solver(solver)
    n_assets = covar.shape[0]
    z = cvx.Variable(n_assets + 1)
    scaled_weights = z[:n_assets]
    exposure_scaler = z[n_assets]
    constraints = [scaled_weights >= 0.0, cvx.sum(scaled_weights) == exposure_scaler]
    if min_weights is not None:
        constraints.append(scaled_weights >= exposure_scaler * min_weights)
    if max_weights is not None:
        constraints.append(scaled_weights <= exposure_scaler * max_weights)
    constraints.append(means.T @ scaled_weights == 1.0)
    problem = cvx.Problem(cvx.Minimize(cvx.quad_form(scaled_weights, covar)), constraints)
    try:
        problem.solve(verbose=False, solver=solver)
    except cvx.error.SolverError as exc:
        raise ParitySolveError(f"{context}: {solver} failed: {exc}") from exc
    if z.value is None or abs(float(z.value[n_assets])) < 1e-14:
        raise ParitySolveError(
            f"{context}: {solver} returned status {problem.status!r} without a usable solution"
        )
    weights = z.value[:n_assets] / z.value[n_assets]
    return _validate_solution(weights, min_weights, max_weights, context=context)


def _cara_objective(weights: np.ndarray, parameters: Sequence[object]) -> float:
    """Expected exponential-CARA disutility across mixture components."""

    means, covariances, probabilities, carra = parameters
    value = 0.0
    for probability, mean, covariance in zip(probabilities, means, covariances):
        value += probability * np.exp(
            -carra * mean.T @ weights
            + 0.5 * carra * carra * weights.T @ covariance @ weights
        )
    return float(value)


def _solve_cara_2024(
    parameters: MixtureParameters,
    carra: float,
    min_weights: np.ndarray | None,
    max_weights: np.ndarray | None,
    context: str,
) -> np.ndarray:
    """Solve one paper-era CARA allocation from equal weights."""

    n_assets = parameters.covariances[0].shape[0]
    # The v2.1.1 code checked Constraints.weights_0 on the class rather than the
    # instance.  Consequently every non-convex CARA solve started equally; that
    # historical behaviour is intentional here.
    initial = np.ones(n_assets) / n_assets
    result = minimize(
        _cara_objective,
        initial,
        args=(
            [
                parameters.means,
                parameters.covariances,
                parameters.probabilities,
                carra,
            ],
        ),
        method="SLSQP",
        constraints=_legacy_scipy_constraints(min_weights, max_weights),
        options={"disp": False, "ftol": 1e-12},
    )
    return _validate_solution(result.x, min_weights, max_weights, context=context)


def _filter_covar(
    covariance: pd.DataFrame,
    vectors: Mapping[str, pd.Series] | None = None,
) -> tuple[pd.DataFrame, Mapping[str, pd.Series] | None]:
    """Apply the public filter with the v2.1.1 default vector treatment."""

    return opt.filter_covar_and_vectors_for_nans(
        pd_covar=covariance,
        vectors=None if vectors is None else dict(vectors),
    )


def _scaled_policy(
    all_columns: pd.Index,
    good_columns: pd.Index,
    min_weights: pd.Series | None,
    max_weights: pd.Series | None,
) -> tuple[np.ndarray | None, np.ndarray | None]:
    """Apply the historical total-to-good scaling to bound vectors."""

    ratio = len(all_columns) / len(good_columns)
    minimum = None
    maximum = None
    if min_weights is not None:
        minimum = min_weights.reindex(good_columns).fillna(0.0).to_numpy()
    if max_weights is not None:
        maximum = ratio * max_weights.reindex(good_columns).fillna(0.0).to_numpy()
    return minimum, maximum


def _report_slice(weights: pd.DataFrame, config: Parity2024Config) -> pd.DataFrame:
    """Restrict target weights to the paper reporting interval."""

    start = _timestamp(config.reporting_start)
    end = _timestamp(config.end_date)
    if start is not None:
        weights = weights.loc[start:]
    if end is not None:
        weights = weights.loc[:end]
    return weights


def _rolling_erc_2024(
    prices: pd.DataFrame,
    risk_budget: pd.Series,
    config: Parity2024Config,
) -> pd.DataFrame:
    """Compute the v2.1.1 ERC target-weight panel."""

    columns = prices.columns
    reporting_start = _timestamp(config.reporting_start)
    weights: dict[pd.Timestamp, pd.Series] = {}
    for date, covariance in estimate_ewm_covariances_2024(prices, config).items():
        if reporting_start is not None and date < reporting_start:
            continue
        clean, _ = _filter_covar(covariance)
        if clean.empty:
            continue
        ratio = len(columns) / len(clean.columns)
        budget = ratio * risk_budget.reindex(clean.columns).fillna(0.0).to_numpy()
        solved = _solve_erc_2024(clean.to_numpy(), budget, context=f"ERC {date.date()}")
        weights[date] = pd.Series(solved, index=clean.columns).reindex(columns).fillna(0.0)
    panel = pd.DataFrame.from_dict(weights, orient="index").reindex(columns=columns)
    return _report_slice(panel, config)


def _rolling_max_div_2024(
    prices: pd.DataFrame,
    min_weights: pd.Series | None,
    max_weights: pd.Series | None,
    config: Parity2024Config,
) -> pd.DataFrame:
    """Compute the v2.1.1 MaxDiv target-weight panel."""

    columns = prices.columns
    reporting_start = _timestamp(config.reporting_start)
    weights: dict[pd.Timestamp, pd.Series] = {}
    for date, covariance in estimate_ewm_covariances_2024(prices, config).items():
        if reporting_start is not None and date < reporting_start:
            continue
        clean, _ = _filter_covar(covariance)
        if clean.empty:
            continue
        minimum, maximum = _scaled_policy(
            columns, clean.columns, min_weights, max_weights
        )
        solved = _solve_max_div_2024(
            clean.to_numpy(), minimum, maximum, context=f"MaxDiv {date.date()}"
        )
        weights[date] = pd.Series(solved, index=clean.columns).reindex(columns).fillna(0.0)
    panel = pd.DataFrame.from_dict(weights, orient="index").reindex(columns=columns)
    return _report_slice(panel, config)


def _rolling_max_sharpe_2024(
    prices: pd.DataFrame,
    min_weights: pd.Series | None,
    max_weights: pd.Series | None,
    config: Parity2024Config,
) -> pd.DataFrame:
    """Compute the v2.1.1 MaxSharpe target-weight panel."""

    inputs = estimate_max_sharpe_inputs_2024(prices, config)
    columns = prices.columns
    reporting_start = _timestamp(config.reporting_start)
    weights: dict[pd.Timestamp, pd.Series] = {}
    for date, covariance in inputs.covariances.items():
        if reporting_start is not None and date < reporting_start:
            continue
        means = inputs.expected_returns.loc[date]
        clean, vectors = _filter_covar(covariance, vectors={"means": means})
        minimum, maximum = _scaled_policy(
            columns, clean.columns, min_weights, max_weights
        )
        solved = _solve_max_sharpe_2024(
            clean.to_numpy(),
            vectors["means"].to_numpy(),
            minimum,
            maximum,
            solver=config.max_sharpe_solver,
            context=f"MaxSharpe {date.date()}",
        )
        weights[date] = pd.Series(solved, index=clean.columns).reindex(columns).fillna(0.0)
    panel = pd.DataFrame.from_dict(weights, orient="index").reindex(columns=columns)
    return _report_slice(panel, config)


def _rolling_cara_2024(
    prices: pd.DataFrame,
    min_weights: pd.Series | None,
    max_weights: pd.Series | None,
    config: Parity2024Config,
    mixture_fitter: MixtureFitter | None,
) -> pd.DataFrame:
    """Compute the v2.1.1 three-component Gaussian-mixture CARA panel."""

    if mixture_fitter is None:
        # Check eagerly so a missing optional dependency never looks like an
        # empty-but-successful result on a short input panel.
        require_parity_dependencies((ParityMethod.CARA,))
        mixture_fitter = fit_sklearn_mixture_2024
    windows = rolling_return_windows_2024(prices, config)
    annualisation = qis.get_annualization_factor(freq=config.returns_freq)
    columns = prices.columns
    start = _timestamp(config.estimation_start)
    reporting_start = _timestamp(config.reporting_start)
    end = _timestamp(config.end_date)
    weights: dict[pd.Timestamp, pd.Series] = {}
    for date, window in windows.items():
        if start is not None and date < start:
            continue
        if reporting_start is not None and date < reporting_start:
            continue
        if end is not None and date > end:
            continue
        clean_returns = window.dropna(axis=1, how="any")
        if clean_returns.shape[1] == 0:
            raise ParitySolveError(f"CARA-3 {date.date()}: no complete assets in the window")
        parameters = mixture_fitter(
            clean_returns.to_numpy(), config.n_components, annualisation
        )
        minimum, maximum = _scaled_policy(
            columns, clean_returns.columns, min_weights, max_weights
        )
        solved = _solve_cara_2024(
            parameters,
            config.carra,
            minimum,
            maximum,
            context=f"CARA-3 {date.date()}",
        )
        weights[date] = (
            pd.Series(solved, index=clean_returns.columns).reindex(columns).fillna(0.0)
        )
    panel = pd.DataFrame.from_dict(weights, orient="index").reindex(columns=columns)
    return _report_slice(panel, config)


def _universe_policy(
    columns: pd.Index,
    is_alternatives: bool,
    first_asset_target_weight: float,
) -> tuple[pd.Series | None, pd.Series | None, pd.Series]:
    """Return the paper's bounds and ERC risk-budget vector for one universe."""

    n_assets = len(columns)
    if n_assets < 2:
        raise ValueError("each marginal universe must contain at least two assets")
    if is_alternatives:
        budget = pd.Series(1.0 / n_assets, index=columns)
        return None, None, budget

    minimum = pd.Series(0.0, index=columns)
    maximum = pd.Series(1.0, index=columns)
    minimum.iloc[0] = first_asset_target_weight
    maximum.iloc[0] = first_asset_target_weight

    # For ERC this 75/25 split is a risk budget, not a 75% capital constraint.
    budget = pd.Series(
        (1.0 - first_asset_target_weight) / (n_assets - 1),
        index=columns,
    )
    budget.iloc[0] = first_asset_target_weight
    return minimum, maximum, budget


def _single_universe_weights_2024(
    prices: pd.DataFrame,
    method: ParityMethod,
    is_alternatives: bool,
    config: Parity2024Config,
    mixture_fitter: MixtureFitter | None,
) -> pd.DataFrame:
    """Run one investable universe under one frozen paper method."""

    prices = _prices_through_end(prices, config)
    minimum, maximum, risk_budget = _universe_policy(
        prices.columns,
        is_alternatives=is_alternatives,
        first_asset_target_weight=config.first_asset_target_weight,
    )
    if method is ParityMethod.ERC:
        return _rolling_erc_2024(prices, risk_budget, config)
    if method is ParityMethod.MAX_DIV:
        return _rolling_max_div_2024(prices, minimum, maximum, config)
    if method is ParityMethod.MAX_SHARPE:
        return _rolling_max_sharpe_2024(prices, minimum, maximum, config)
    if method is ParityMethod.CARA:
        return _rolling_cara_2024(
            prices, minimum, maximum, config, mixture_fitter=mixture_fitter
        )
    raise AssertionError(f"Unhandled method {method}")


def _validate_weight_panel(
    weights: pd.DataFrame,
    expected_columns: pd.Index,
    label: str,
    tolerance: float = 2e-5,
) -> None:
    """Verify labels, finiteness, full investment and long-only semantics."""

    if weights.empty:
        raise ParitySolveError(f"{label}: no eligible target weights")
    if not weights.columns.equals(expected_columns):
        raise ParitySolveError(f"{label}: weight labels differ from the investable universe")
    values = weights.to_numpy(dtype=float)
    if not np.all(np.isfinite(values)):
        raise ParitySolveError(f"{label}: weights contain non-finite values")
    if float(np.min(values)) < -tolerance:
        raise ParitySolveError(f"{label}: negative long-only weight")
    error = float(np.max(np.abs(weights.sum(axis=1).to_numpy() - 1.0)))
    if error > tolerance:
        raise ParitySolveError(f"{label}: maximum weight-sum error is {error:.3g}")


def marginal_weights_2024(
    prices: pd.DataFrame,
    marginal_asset: str,
    method: ParityMethod | str | object,
    is_alternatives: bool,
    config: Parity2024Config = Parity2024Config(),
    mixture_fitter: MixtureFitter | None = None,
) -> MarginalWeights:
    """Run separately estimated without/with-asset paper target weights.

    Separate calls are intentional and load-bearing.  Passing the with-asset
    covariance to the without-asset MaxDiv problem, as the current shared
    marginal wrapper does, drops the crypto allocation after optimisation and
    leaves implicit cash.
    """

    prices_with = _prices_through_end(prices, config)
    if marginal_asset not in prices_with.columns:
        raise KeyError(f"Missing marginal asset {marginal_asset!r}")
    prices_without = prices_with.drop(columns=marginal_asset)
    method = _coerce_method(method)

    without = _single_universe_weights_2024(
        prices_without,
        method=method,
        is_alternatives=is_alternatives,
        config=config,
        mixture_fitter=mixture_fitter,
    )
    with_asset = _single_universe_weights_2024(
        prices_with,
        method=method,
        is_alternatives=is_alternatives,
        config=config,
        mixture_fitter=mixture_fitter,
    )
    _validate_weight_panel(without, prices_without.columns, f"{method.value} without")
    _validate_weight_panel(with_asset, prices_with.columns, f"{method.value} with")
    return MarginalWeights(without_asset=without, with_asset=with_asset)


def backtest_marginal_2024(
    prices: pd.DataFrame,
    marginal_asset: str,
    method: ParityMethod | str | object,
    is_alternatives: bool,
    config: Parity2024Config = Parity2024Config(),
    mixture_fitter: MixtureFitter | None = None,
) -> MarginalPortfolios:
    """Backtest one historical with/without pair using the 2024 lag and costs."""

    prices_with = _prices_through_end(prices, config)
    prices_without = prices_with.drop(columns=marginal_asset)
    method = _coerce_method(method)
    weights = marginal_weights_2024(
        prices=prices_with,
        marginal_asset=marginal_asset,
        method=method,
        is_alternatives=is_alternatives,
        config=config,
        mixture_fitter=mixture_fitter,
    )
    without = qis.backtest_model_portfolio(
        prices=qis.truncate_prior_to_start(
            prices_without, start=weights.without_asset.index[0]
        ),
        weights=weights.without_asset,
        rebalancing_freq=config.rebalancing_freq,
        is_rebalanced_at_first_date=True,
        rebalancing_costs=config.rebalancing_costs,
        weight_implementation_lag=config.weight_implementation_lag,
        ticker=f"{method.value} w/o {marginal_asset}",
    )
    with_asset = qis.backtest_model_portfolio(
        prices=qis.truncate_prior_to_start(
            prices_with, start=weights.with_asset.index[0]
        ),
        weights=weights.with_asset,
        rebalancing_freq=config.rebalancing_freq,
        is_rebalanced_at_first_date=True,
        rebalancing_costs=config.rebalancing_costs,
        weight_implementation_lag=config.weight_implementation_lag,
        ticker=f"{method.value} with {marginal_asset}",
    )
    return MarginalPortfolios(
        without_asset=without,
        with_asset=with_asset,
        weights=weights,
    )


def golden_headline_from_constants() -> float:
    """Return the median of the four archived cross-method scenario medians."""

    scenario_medians = [row[-1] for row in GOLDEN_MEDIAN_WEIGHT_ROWS.values()]
    return float(np.median(scenario_medians))


def verify_golden_workbook(
    path: Path = GOLDEN_WORKBOOK_PATH,
    tolerance: float = 1e-12,
) -> Mapping[str, tuple[float, ...]]:
    """Verify the archived workbook checksum and its B3:F3 oracle cells."""

    if not path.is_file():
        raise FileNotFoundError(path)
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    if digest.lower() != GOLDEN_WORKBOOK_SHA256:
        raise AssertionError(f"Unexpected golden workbook SHA-256: {digest}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ParityDependencyError("openpyxl is required to inspect the golden workbook") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    observed: dict[str, tuple[float, ...]] = {}
    for sheet, expected in GOLDEN_MEDIAN_WEIGHT_ROWS.items():
        values = tuple(float(workbook[sheet].cell(row=3, column=column).value) for column in range(2, 7))
        if not np.allclose(values, expected, rtol=0.0, atol=tolerance):
            raise AssertionError(f"Golden cells changed in {sheet}: {values!r}")
        observed[sheet] = values
    return observed


def golden_runtime_available() -> tuple[bool, str]:
    """Return whether the optional archived-panel integration replay can run."""

    missing: list[str] = []
    if not GOLDEN_PANEL_PATH.is_file():
        missing.append(f"archived panel is absent: {GOLDEN_PANEL_PATH}")
    if not GOLDEN_WORKBOOK_PATH.is_file():
        missing.append(f"archived workbook is absent: {GOLDEN_WORKBOOK_PATH}")
    report = dependency_report(methods=(ParityMethod.MAX_SHARPE, ParityMethod.CARA))
    missing.extend(report.missing_or_incompatible)
    if _installed_version("openpyxl") is None:
        missing.append("openpyxl is not installed")
    return not missing, "; ".join(missing)
