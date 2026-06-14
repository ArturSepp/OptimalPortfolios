"""
run_exhibits.py — regenerate the paper's figures from cached production inputs.

Cached inputs in ./data/ (built from the production workbook at 31 Mar 2026):
    inputs17_v5.npz : lam, Sig_F, beta, D, mu_alpha, w_b, tickers, labels, FACT
    allfig17.npz    : mu, Sigma, w_b, sr_tan, sat, sat_order, labels
    cascade_data.npz: headline + 8-mandate TE-constrained Sharpe ratios

Figure mapping (paper numbering):
    1  exhibit_eq7_weights          eq7()
    2  exhibit_factor_mimicking     mimicking()
    3  exhibit_sharpe2_waterfall    waterfall()
    4  exhibit_eq11_illustration    per_factor_identity()
    5  exhibit_2_per_factor         per_factor_gap()
    6  exhibit_id_deficit           deficit()
    7  exhibit_candidate_value      candidates()
    8  exhibit_3_scaling            saturation()
    9  exhibit_5_constraint_cost    cascade()
    10 exhibit_bootstrap            run_bootstrap.py
    11 exhibit_tre_grid             tre()
House style: serif, navy #1f3864 / steel #2E6DB4 / orange #E8943A / grey #888888,
no in-figure titles (captions live in the paper).
"""
from __future__ import annotations
import os
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.patches import Patch

import identity_analytics as ia

HERE = os.path.dirname(os.path.abspath(__file__))
DATA, FIG = os.path.join(HERE, "data"), os.path.join(HERE, "figures")
NAVY, STEEL, LT, ORANGE, GREY = '#1f3864', '#2E6DB4', '#9BC4EC', '#E8943A', '#888888'
plt.rcParams.update({'font.family': 'serif', 'font.size': 10})

d = np.load(os.path.join(DATA, 'inputs17_v5.npz'), allow_pickle=True)
a = np.load(os.path.join(DATA, 'allfig17.npz'), allow_pickle=True)
lam, Sig_F, beta, D = d['lam'], d['Sig_F'], d['beta'], d['D']
mu_alpha = d['mu_alpha']
labels, FACT = list(d['labels']), list(d['FACT'])
N, M = beta.shape


def _save(fig, name):
    fig.tight_layout()
    fig.savefig(os.path.join(FIG, name), dpi=160, bbox_inches='tight')
    plt.close(fig)
    print(f"  written {name}")


def eq7():
    ag = np.asarray(a['alpha'])  # GLS-projected alpha: the paper's convention
    w_sys, w_idio = ia.eq7_weight_blocks(lam, Sig_F, beta, D, ag)
    g = (w_sys + w_idio).sum()
    sysN, idiN = 100 * w_sys / g, 100 * w_idio / g
    tot = sysN + idiN
    order = np.argsort(-tot)
    y = np.arange(N)[::-1]
    fig, ax = plt.subplots(figsize=(9, 5.8))
    ax.barh(y, sysN[order], color=STEEL, height=0.66, zorder=2)
    ax.barh(y, idiN[order],
            left=np.where(idiN[order] >= 0, np.maximum(sysN[order], 0),
                          np.minimum(sysN[order], 0)),
            color=ORANGE, height=0.66, zorder=2)
    for yi, tt in zip(y, tot[order]):
        ax.plot([tt, tt], [yi - 0.34, yi + 0.34], color='black', lw=2, zorder=4)
    ax.axvline(0, color='black', lw=0.8)
    ax.set_yticks(y); ax.set_yticklabels([labels[i] for i in order], fontsize=9)
    ax.set_xlabel('Share of Sharpe-maximizing portfolio (%)')
    ax.grid(axis='x', alpha=0.25, zorder=0)
    ax.legend(handles=[
        Patch(color=STEEL, label='factor-premium weight  (systematic, $P_F(\\Sigma_F+\\beta_F^{-1})^{-1}\\lambda$)'),
        Patch(color=ORANGE, label='alpha weight  (idiosyncratic, $D^{-1}\\alpha_{\\mathrm{gls}}$ — factor-neutral overlay)')],
        fontsize=8.5, loc='lower right', framealpha=0.93)
    _save(fig, 'exhibit_eq7_weights.png')


def mimicking():
    P_F = ia.factor_mimicking_portfolios(beta, D)
    fig, ax = plt.subplots(figsize=(8.5, 6.2))
    vmax = np.abs(P_F).max()
    im = ax.imshow(P_F, cmap='RdBu_r', vmin=-vmax, vmax=vmax, aspect='auto')
    ax.set_xticks(range(M)); ax.set_xticklabels(FACT, rotation=35, ha='right', fontsize=8.5)
    ax.set_yticks(range(N)); ax.set_yticklabels(labels, fontsize=8.5)
    fig.colorbar(im, ax=ax, shrink=0.8, label='$P_F$ weight')
    _save(fig, 'exhibit_factor_mimicking.png')


def waterfall():
    ceil = ia.frictionless_ceiling(lam, Sig_F)
    ident = ia.matf_identity(lam, Sig_F, beta, D)
    a_gls = ia.gls_orthogonal_alpha(mu_alpha, beta, D)
    sr2a = ia.idiosyncratic_sr2(a_gls, D)
    total = ident + sr2a
    fig, ax = plt.subplots(figsize=(8, 4.4))
    x = np.arange(4)
    ax.bar(0, ident, color=NAVY, width=0.62)
    ax.bar(0, ceil - ident, bottom=ident, color=NAVY, width=0.62,
           hatch='//', edgecolor='white')
    ax.bar(1, ident, color=STEEL, width=0.62)
    ax.bar(2, total, color=ORANGE, width=0.62)
    ax.bar(3, total, color='#555555', width=0.62)
    for xi, v in zip(x, [ceil, ident, total, total]):
        ax.text(xi, v + 0.012, f'{v:.2f}', ha='center', fontsize=9)
    ax.set_xticks(x)
    ax.set_xticklabels(['Frictionless\nceiling', 'MATF\nIdentity',
                        '+ idiosyncratic\n(ILS, RA)', 'Total'], fontsize=9)
    ax.set_ylabel('Squared Sharpe')
    ax.grid(axis='y', alpha=0.25)
    _save(fig, 'exhibit_sharpe2_waterfall.png')


def per_factor_identity():
    c_id, _ = ia.per_factor_contributions(lam, Sig_F, beta, D)
    bF = ia.factor_information_matrix(beta, D)
    u = bF @ np.linalg.solve(np.eye(M) + Sig_F @ bF, lam)
    order = np.argsort(-c_id)
    x = np.arange(M)
    fig, ax = plt.subplots(figsize=(9.2, 4.6))
    ax.bar(x, c_id[order], color=STEEL, width=0.62, zorder=3)
    for xi, i in zip(x, order):
        ax.text(xi, c_id[i] + (0.004 if c_id[i] >= 0 else -0.012),
                f'$\\lambda$={100*lam[i]:.2f}%\n$u$={u[i]:.2f}',
                ha='center', fontsize=7.2)
    ax.axhline(0, color='black', lw=0.8)
    ax.set_xticks(x); ax.set_xticklabels([FACT[i] for i in order],
                                         rotation=35, ha='right', fontsize=9)
    ax.set_ylabel('Per-factor SR$^2$ contribution')
    ax.text(0.98, 0.95, f'$\\Sigma$ = {c_id.sum():.2f}  (SR = {np.sqrt(c_id.sum()):.2f})',
            transform=ax.transAxes, ha='right', fontsize=9,
            bbox=dict(facecolor='white', alpha=0.9, edgecolor=GREY))
    ax.grid(axis='y', alpha=0.25, zorder=0)
    _save(fig, 'exhibit_eq11_illustration.png')


def per_factor_gap():
    c_id, c_ce = ia.per_factor_contributions(lam, Sig_F, beta, D)
    order = np.argsort(-c_ce)
    x = np.arange(M); w = 0.38
    fig, ax = plt.subplots(figsize=(9.2, 4.6))
    ax.bar(x - w / 2, c_ce[order], w, color=GREY,
           label='frictionless ceiling contribution')
    ax.bar(x + w / 2, c_id[order], w, color=STEEL,
           label='MATF Identity contribution')
    ax.axhline(0, color='black', lw=0.8)
    ax.set_xticks(x); ax.set_xticklabels([FACT[i] for i in order],
                                         rotation=35, ha='right', fontsize=9)
    ax.set_ylabel('Per-factor SR$^2$ contribution')
    ax.legend(fontsize=9); ax.grid(axis='y', alpha=0.25)
    _save(fig, 'exhibit_2_per_factor.png')


def deficit():
    c_id, c_ce = ia.per_factor_contributions(lam, Sig_F, beta, D)
    gap = c_ce - c_id
    share = 100 * gap / gap.sum()
    dfc = ia.identification_deficit(lam, beta, D)
    order = sorted(range(M), key=lambda k: -share[k])
    x = np.arange(M)
    fig, axes = plt.subplots(1, 2, figsize=(10, 4.4))
    axA, axB = axes
    for xi, k in zip(x, order):
        if np.isfinite(dfc[k]):
            axA.bar(xi, dfc[k], color=STEEL, width=0.62, zorder=3)
        else:
            axA.text(xi, 0.3, 'n/a', ha='center', fontsize=8, color=GREY,
                     rotation=90)
    axA.set_title('(A) Identification deficit  $\\mathrm{SE}_k/\\lambda_k$',
                  fontsize=10)
    for xi, k in zip(x, order):
        axB.bar(xi, share[k], color=ORANGE if share[k] >= 0 else GREY,
                width=0.62, zorder=3)
    axB.axhline(0, color='black', lw=0.8)
    axB.set_title('(B) Share of ceiling-to-identity gap (%)', fontsize=10)
    for ax in axes:
        ax.set_xticks(x)
        ax.set_xticklabels([FACT[k] for k in order], rotation=35,
                           ha='right', fontsize=8.5)
        ax.grid(axis='y', alpha=0.25, zorder=0)
    _save(fig, 'exhibit_id_deficit.png')


CANDIDATES = [('Rates-vol (swaption)', 'Rates Vol', 0.05),
              ('CDX credit index', 'Credit', 0.04),
              ('Carry basket', 'Carry', 0.05),
              ('PE-replication sleeve', 'Private Equity', 0.08),
              ('Commodity index', 'Commodities', 0.10),
              ('TIPS / inflation swap', 'Inflation', 0.04),
              ('Duration swap', 'Rates', 0.03),
              ('Equity-beta ETF', 'Equity', 0.15)]


def candidates():
    rows = []
    for name, fac, rv in CANDIDATES:
        bnew = np.zeros(M); bnew[FACT.index(fac)] = 1.0
        rows.append((name, 100 * ia.candidate_fpir_gain(lam, Sig_F, beta, D,
                                                        bnew, rv ** 2)))
    rows.sort(key=lambda r: -r[1])
    y = np.arange(len(rows))[::-1]
    fig, ax = plt.subplots(figsize=(8.5, 4.8))
    ax.barh(y, [r[1] for r in rows], color=STEEL, height=0.62, zorder=3)
    for yi, (_, v) in zip(y, rows):
        ax.text(v + 0.08, yi, f'+{v:.1f}pp', va='center', fontsize=8.5)
    ax.set_yticks(y); ax.set_yticklabels([r[0] for r in rows], fontsize=9)
    ax.set_xlabel('Marginal FPIR gain from adding each candidate alone (percentage points)')
    ax.set_xlim(0, max(r[1] for r in rows) * 1.18)
    ax.grid(axis='x', alpha=0.25, zorder=0)
    _save(fig, 'exhibit_candidate_value.png')


def saturation():
    sat, order = a['sat'], a['sat_order']
    fig, ax = plt.subplots(figsize=(8.5, 4.4))
    ax.plot(order, sat, 'o-', color=NAVY, lw=1.6, ms=4.5, zorder=3)
    ax.axhline(ia.frictionless_ceiling(lam, Sig_F), color=GREY, ls='--', lw=1.2,
               label='frictionless ceiling')
    ax.set_xlabel('Number of assets $N$ (greedy identification ordering)')
    ax.set_ylabel('MATF Identity SR$^2$')
    ax.legend(fontsize=9); ax.grid(alpha=0.25, zorder=0)
    _save(fig, 'exhibit_3_scaling.png')


def cascade():
    c = np.load(os.path.join(DATA, 'cascade_data.npz'), allow_pickle=True)
    scl, sid, slo = float(c['sr_ceiling']), float(c['sr_identity']), float(c['sr_lo'])
    mandates, srs = list(c['mandates']), list(c['srs'])
    vals = [scl, sid, slo] + [float(s) for s in srs]
    labs = ['Frictionless\nceiling', 'MATF\nIdentity', 'Long-only\ntangency'] + \
           [m.replace(' w/o ', '\nw/o ').replace(' with ', '\nwith ') for m in mandates]
    cols = [NAVY, STEEL, GREY] + [ORANGE if 'with' in m else GREY for m in mandates]
    x = np.arange(len(vals))
    fig, ax = plt.subplots(figsize=(10, 5.2))
    ax.bar(x, vals, color=cols, zorder=3, width=0.68)
    for xi, v in zip(x, vals):
        ax.text(xi, v + 0.012, f'{v:.2f}', ha='center', fontsize=8.3)
    for yv, cc in [(scl, NAVY), (sid, STEEL), (slo, GREY)]:
        ax.axhline(yv, ls='--', color=cc, lw=0.9, alpha=0.55, zorder=1)
    ax.set_xticks(x); ax.set_xticklabels(labs, fontsize=8, rotation=35, ha='right')
    ax.set_ylabel('Sharpe Ratio'); ax.set_ylim(0, scl * 1.12)
    ax.grid(axis='y', alpha=0.25, zorder=0)
    ax.axvline(2.5, color='black', lw=0.6, alpha=0.4)
    _save(fig, 'exhibit_5_constraint_cost.png')


def tre():
    mu, Sigma, w_b = a['mu'], a['Sigma'], a['w_b']
    labs = list(a['labels']); sr_tan = float(a['sr_tan'])
    TEs = [0.01, 0.02, 0.03]; cols = {0.01: LT, 0.02: STEEL, 0.03: NAVY}
    order = np.argsort(-w_b); y = np.arange(len(labs))[::-1]; bw = 0.2
    fig, ax = plt.subplots(figsize=(9, 6.3))
    ax.barh(y + 1.5 * bw, 100 * w_b[order], height=bw, color=GREY,
            label='Benchmark', zorder=3)
    for k, te in enumerate(TEs):
        dw, _ = ia.te_overlay(mu, Sigma, te)
        ax.barh(y + (0.5 - k) * bw, 100 * (w_b + dw)[order], height=bw,
                color=cols[te], label=f'TE = {int(100 * te)}%', zorder=3)
    ax.axvline(0, color='black', lw=0.8)
    ax.set_yticks(y); ax.set_yticklabels([labs[i] for i in order], fontsize=9)
    ax.set_xlabel('Weight (%)'); ax.set_xlim(-32, 46)
    ax.legend(fontsize=8.5, loc='lower right'); ax.grid(axis='x', alpha=0.25, zorder=0)
    ax.text(0.02, 0.97, f'IR = {sr_tan:.2f} at every TE budget',
            transform=ax.transAxes, fontsize=9.5, style='italic', va='top')
    _save(fig, 'exhibit_tre_grid.png')


def alpha_projection():
    """Admitted vs GLS-projected alpha, per asset (GLS-projection exhibit)."""
    from matplotlib.ticker import FuncFormatter
    aa = 100 * mu_alpha
    ag = 100 * np.asarray(a['alpha'])
    y = np.arange(N)[::-1]
    bw = 0.38
    fig, ax = plt.subplots(figsize=(9, 5.8))
    ax.barh(y + bw / 2 + 0.02, aa, height=bw, color=STEEL, zorder=3)
    ax.barh(y - bw / 2 - 0.02, ag, height=bw, color=ORANGE, zorder=3)
    ax.axvline(0, color='black', lw=0.8)
    ax.set_yticks(y); ax.set_yticklabels(labels, fontsize=9)
    ax.set_xlabel('Alpha (% per annum)')
    ax.xaxis.set_major_formatter(FuncFormatter(lambda v, p: f'{v:.1f}%'))
    ax.grid(axis='x', alpha=0.25, zorder=0)
    _save(fig, 'exhibit_alpha_projection.png')


def universe_table():
    """Appendix exhibit: full 17-asset universe snapshot (betas, R2, alpha, CMA, vol)."""
    import openpyxl, paper_inputs as pi
    wb = openpyxl.load_workbook(pi.XLSX, data_only=True)
    m = wb['cma_metadata']
    hdr = {m.cell(1, c).value: c for c in range(1, m.max_column + 1)}
    mrow = {m.cell(r, 1).value: r for r in range(2, m.max_row + 1)}
    v = wb['y_variances']
    vrow = {v.cell(r, 1).value: r for r in range(2, v.max_row + 1)}
    rows = []
    for t, lab in zip(pi.TICKERS, labels):
        key = f"{t} Index"
        g = lambda col: float(m.cell(mrow[key], hdr[col]).value or 0)
        rows.append(dict(label=lab, ticker=t.replace('_', chr(92) + '_'),
                         ac=m.cell(mrow[key], hdr['asset_class']).value,
                         b=beta[labels.index(lab)],
                         r2=float(v.cell(vrow[key], 5).value or 0),
                         alpha=g('excess_alpha_cma'), cma=g('base_excess_cma'),
                         vol=g('total_vol')))
    cols = ['Asset', 'Ticker'] + [fc[:3].upper() if fc != 'Private Equity' else 'PE'
                                  for fc in FACT] + ['$R^2$', 'Alpha', 'CMA', 'Vol']
    cols = ['Asset', 'Ticker', 'EQ', 'RAT', 'CRD', 'CRY', 'INF', 'COM', 'PE', 'RVOL', 'FX',
            'R2', 'Alpha', 'CMA', 'Vol']
    fig, ax = plt.subplots(figsize=(13.2, 6.4))
    ax.axis('off')
    nr, nc = len(rows) + 1, len(cols)
    wpos = np.array([0.16, 0.10] + [0.052] * 9 + [0.052, 0.06, 0.06, 0.055])
    x = np.concatenate([[0], np.cumsum(wpos)]) / wpos.sum()
    yy = np.linspace(1, 0, nr + 1)
    for j, c in enumerate(cols):
        ax.text((x[j] + x[j + 1]) / 2, (yy[0] + yy[1]) / 2, c, ha='center', va='center',
                fontsize=8.6, fontweight='bold', color='white',
                bbox=None, transform=ax.transAxes)
    ax.add_patch(plt.Rectangle((0, yy[1]), 1, yy[0] - yy[1], color=NAVY,
                               transform=ax.transAxes, zorder=0))
    prev_ac = None
    for i, r in enumerate(rows):
        yt = (yy[i + 1] + yy[i + 2]) / 2
        if i % 2 == 1:
            ax.add_patch(plt.Rectangle((0, yy[i + 2]), 1, yy[i + 1] - yy[i + 2],
                                       color='#EDF2F9', transform=ax.transAxes, zorder=0))
        if r['ac'] != prev_ac:
            ax.plot([0, 1], [yy[i + 1], yy[i + 1]], color=NAVY, lw=1.2,
                    transform=ax.transAxes, zorder=1)
            prev_ac = r['ac']
        vals = [r['label'], r['ticker'].replace(chr(92) + '_', '_')]
        vals += ['' if abs(b) < 5e-3 else f"{b:.2f}" for b in r['b']]
        vals += [f"{100 * r['r2']:.0f}%",
                 '' if abs(r['alpha']) < 5e-5 else f"{100 * r['alpha']:.2f}%",
                 f"{100 * r['cma']:.2f}%", f"{100 * r['vol']:.1f}%"]
        for j, val in enumerate(vals):
            ha = 'left' if j == 0 else ('left' if j == 1 else 'center')
            xt = x[j] + 0.004 if ha == 'left' else (x[j] + x[j + 1]) / 2
            ax.text(xt, yt, val, ha=ha, va='center', fontsize=8.2,
                    transform=ax.transAxes)
    ax.plot([0, 1], [yy[-1], yy[-1]], color=NAVY, lw=1.2, transform=ax.transAxes)
    _save(fig, 'exhibit_universe_snapshot.png')


ALL = [eq7, mimicking, waterfall, per_factor_identity, per_factor_gap,
       deficit, candidates, cascade, tre, alpha_projection, universe_table]

if __name__ == '__main__':
    print("[exhibits] regenerating from cached production inputs...")
    for fn in ALL:
        fn()
    print("[exhibits] done; bootstrap figure via run_bootstrap.py")
