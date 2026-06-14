"""
paper_inputs.py — load production inputs and build bootstrap panels.

Data files expected in ./data/:
    global_saa_universe_data_cmas_usd_newbeta_2026q1.xlsx   (production CMA workbook)
    futures_risk_factors.csv                                (daily factor NAVs, base 100)
    universe_snapshot.xlsx                                  (mandate benchmark weights)

Conventions (production workbook, 2026Q1 vintage):
    factor_cmas  col 2          : lambda (9,)
    x_covar      9x9            : Sig_F, annualized
    y_betas      cols 2..10     : beta (per asset row, '<TICKER> Index')
    y_variances  col 3 / col 5  : residual variance / R^2
    cma_metadata col 21 / 72    : alpha admission weight / admitted excess alpha
"""
from __future__ import annotations
import os
import numpy as np
import pandas as pd
import openpyxl

DATA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
XLSX = os.path.join(DATA, "global_saa_universe_data_cmas_usd_newbeta_2026q1.xlsx")
FRF = os.path.join(DATA, "futures_risk_factors.csv")

TICKERS = ['LGTRTRUH', 'LGCPTRUH', 'H23059US', 'EMUSTRUU', 'LF94TRUH',
           'NDDUUS', 'MSDEE15N', 'NDDLJN', 'NDDLUK', 'SLIC', 'M1APJ',
           'M1EFZ', 'MP503001', 'MP503008', 'LGT_ILS', 'HFRXGL',
           'LGT_REAL_ASSETS']
QE_ASSETS = ('MP503001', 'MP503008', 'LGT_ILS')
FACTORS = ['Equity', 'Rates', 'Credit', 'Carry', 'Inflation', 'Commodities',
           'Private Equity', 'Rates Vol', 'Fx']


def load_production_inputs(xlsx_path: str = XLSX) -> dict:
    """Production (lam, Sig_F, beta, D, alpha) for the 17-asset universe."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    lam = np.array([wb['factor_cmas'].cell(r, 2).value for r in range(2, 11)], float)
    Sig_F = np.array([[float(wb['x_covar'].cell(r, c).value)
                       for c in range(2, 11)] for r in range(2, 11)])
    m, b, v = wb['cma_metadata'], wb['y_betas'], wb['y_variances']
    mrow = {m.cell(r, 1).value: r for r in range(2, m.max_row + 1)}
    brow = {b.cell(r, 1).value: r for r in range(2, b.max_row + 1)}
    vrow = {v.cell(r, 1).value: r for r in range(2, v.max_row + 1)}
    beta = np.array([[b.cell(brow[f"{t} Index"], c).value for c in range(2, 11)]
                     for t in TICKERS], float)
    D = np.array([float(v.cell(vrow[f"{t} Index"], 3).value) for t in TICKERS])
    alpha = np.array([float(m.cell(mrow[f"{t} Index"], 72).value or 0.0)
                      for t in TICKERS])
    return dict(lam=lam, Sig_F=Sig_F, beta=beta, D=D, alpha=alpha,
                tickers=TICKERS, factors=FACTORS)


# Bootstrap window: Apr 2001 - Mar 2026 (T = 300 months), identical to the
# panel window of the MATF-CMA bootstrap (matf_cma_jpm_2026, paper Appendix B).
BOOT_START = "2001-04-01"
BOOT_END = "2026-03-31"


def build_bootstrap_panels(inputs: dict | None = None,
                           frf_path: str = FRF,
                           xlsx_path: str = XLSX):
    """Paired bootstrap panels: monthly factor log returns F (T x 9) and
    native-frequency residuals E (T x 17), residuals computed with the
    production beta (quarterly factor sums for QE assets). Panels are
    trimmed to BOOT_START..BOOT_END to match the MATF-CMA bootstrap."""
    inputs = inputs or load_production_inputs(xlsx_path)
    frf = pd.read_csv(frf_path, index_col=0, parse_dates=True)
    F = np.log(frf.resample('ME').last()).diff().dropna()
    F = F[FACTORS] if list(F.columns) != FACTORS else F

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['excess_logreturns']
    hdr = [ws.cell(1, c).value for c in range(2, ws.max_column + 1)]
    dates = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
    vals = [[ws.cell(r, c).value for c in range(2, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1)]
    A = pd.DataFrame(vals, index=pd.to_datetime(dates), columns=hdr).astype(float)
    A = A[[c for c in A.columns if str(c).replace(' Index', '') in TICKERS]]
    A.columns = [str(c).replace(' Index', '') for c in A.columns]
    A = A[TICKERS]

    common = F.index.intersection(A.index)
    common = common[(common >= BOOT_START) & (common <= BOOT_END)]
    F, A = F.loc[common], A.loc[common]
    F3 = F.rolling(3).sum()
    E = pd.DataFrame(index=common, columns=TICKERS, dtype=float)
    for j, t in enumerate(TICKERS):
        bvec = inputs['beta'][j]
        E[t] = A[t] - (F3 @ bvec if t in QE_ASSETS else F @ bvec)
    return F, E


def fpy_per_asset() -> np.ndarray:
    return np.array([4.0 if t in QE_ASSETS else 12.0 for t in TICKERS])
